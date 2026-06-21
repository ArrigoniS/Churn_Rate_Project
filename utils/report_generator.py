"""
Report Generator — PDF + Excel
Genera report scaricabili con classifica dipendenti a rischio e analisi LLM.
Richiede: pip install reportlab openpyxl
"""

import os
import io
from datetime import datetime
from typing import Optional

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ─────────────────────────────────────────────
# COLORI BRAND
# ─────────────────────────────────────────────
C_PRIMARY    = colors.HexColor("#1a1a2e")
C_ACCENT     = colors.HexColor("#e94560")
C_CRITICO    = colors.HexColor("#dc2626")
C_ALTO       = colors.HexColor("#ea580c")
C_MEDIO      = colors.HexColor("#ca8a04")
C_BASSO      = colors.HexColor("#16a34a")
C_LIGHT_GRAY = colors.HexColor("#f8fafc")
C_BORDER     = colors.HexColor("#e2e8f0")
C_TEXT       = colors.HexColor("#1e293b")
C_MUTED      = colors.HexColor("#64748b")

RISK_COLORS = {
    "Critico": C_CRITICO,
    "Alto":    C_ALTO,
    "Medio":   C_MEDIO,
    "Basso":   C_BASSO,
}

RISK_HEX = {
    "Critico": "DC2626",
    "Alto":    "EA580C",
    "Medio":   "CA8A04",
    "Basso":   "16A34A",
}


# ─────────────────────────────────────────────
# PDF REPORT
# ─────────────────────────────────────────────

def _build_styles():
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        "Title2", parent=styles["Normal"],
        fontSize=24, textColor=C_PRIMARY,
        fontName="Helvetica-Bold", spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=11, textColor=C_MUTED,
        fontName="Helvetica", spaceAfter=20,
    ))
    styles.add(ParagraphStyle(
        "SectionHeader", parent=styles["Normal"],
        fontSize=14, textColor=C_PRIMARY,
        fontName="Helvetica-Bold", spaceBefore=16, spaceAfter=8,
        borderPad=4,
    ))
    styles.add(ParagraphStyle(
        "CardTitle", parent=styles["Normal"],
        fontSize=12, textColor=colors.white,
        fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        "CardBody", parent=styles["Normal"],
        fontSize=9, textColor=C_TEXT,
        fontName="Helvetica", spaceAfter=4, leading=13,
    ))
    styles.add(ParagraphStyle(
        "Label", parent=styles["Normal"],
        fontSize=8, textColor=C_MUTED,
        fontName="Helvetica-Bold",
    ))
    styles.add(ParagraphStyle(
        "Value", parent=styles["Normal"],
        fontSize=9, textColor=C_TEXT,
        fontName="Helvetica",
    ))
    styles.add(ParagraphStyle(
        "Footer", parent=styles["Normal"],
        fontSize=8, textColor=C_MUTED,
        fontName="Helvetica", alignment=TA_CENTER,
    ))
    return styles


def _summary_table(df_scored: pd.DataFrame, styles) -> list:
    """Tabella riepilogativa con conteggi per livello di rischio."""
    elements = []
    elements.append(Paragraph("Riepilogo per livello di rischio", styles["SectionHeader"]))

    risk_counts = df_scored["risk_level"].value_counts()
    total = len(df_scored)

    data = [["Livello", "Dipendenti", "% sul totale", "Score medio"]]
    for level in ["Critico", "Alto", "Medio", "Basso"]:
        count = risk_counts.get(level, 0)
        pct = count / total * 100 if total > 0 else 0
        avg_score = df_scored[df_scored["risk_level"] == level]["churn_score"].mean()
        avg_score_str = f"{avg_score:.1f}" if count > 0 else "—"
        data.append([level, str(count), f"{pct:.1f}%", avg_score_str])

    data.append(["TOTALE", str(total), "100%", f"{df_scored['churn_score'].mean():.1f}"])

    col_widths = [4*cm, 3.5*cm, 4*cm, 4*cm]
    t = Table(data, colWidths=col_widths)

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), C_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [C_LIGHT_GRAY, colors.white]),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, C_BORDER),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f1f5f9")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ])

    # Colora la colonna Livello con i colori di rischio
    for i, level in enumerate(["Critico", "Alto", "Medio", "Basso"], start=1):
        style.add("TEXTCOLOR", (0, i), (0, i), RISK_COLORS.get(level, C_TEXT))
        style.add("FONTNAME", (0, i), (0, i), "Helvetica-Bold")

    t.setStyle(style)
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))
    return elements


def _employee_card(employee: dict, analysis: dict, styles, index: int) -> list:
    """Genera la scheda PDF per un singolo dipendente."""
    elements = []

    risk = str(employee.get("risk_level", "Medio"))
    score = employee.get("churn_score", 0)
    risk_color = RISK_COLORS.get(risk, C_MEDIO)
    nome_completo = f"{employee.get('nome', '')} {employee.get('cognome', '')}"

    # Header card colorato
    header_data = [[
        Paragraph(f"#{index}  {nome_completo}", styles["CardTitle"]),
        Paragraph(
            f"<b>{risk.upper()}</b>  —  Score {score:.0f}/100",
            ParagraphStyle("RiskBadge", parent=styles["CardTitle"],
                           alignment=TA_RIGHT, textColor=colors.white)
        ),
    ]]
    header_table = Table(header_data, colWidths=[11*cm, 6.5*cm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), risk_color),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (0, -1), 12),
        ("RIGHTPADDING", (-1, 0), (-1, -1), 12),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROUNDEDCORNERS", [4, 4, 0, 0]),
    ]))
    elements.append(header_table)

    # Body card con dati e analisi
    dip = employee
    info_rows = [
        ["Dipartimento", dip.get("dipartimento", "N/D"),
         "Ruolo", dip.get("ruolo", "N/D")],
        ["Età", f"{dip.get('eta', 'N/D')} anni",
         "Anzianità", f"{dip.get('anni_in_azienda', 'N/D'):.1f} anni"],
        ["Stipendio annuo", f"€{dip.get('stipendio_annuo', 0):,}",
         "Crescita stipendio", f"{dip.get('crescita_stipendio_percentuale', 'N/D')}% annua"],
        ["Ore sett. medie", f"{dip.get('ore_settimanali_medie', 'N/D')}h",
         "Ultimo check-in HR", f"{dip.get('giorni_ultimo_checkin_hr', 'N/D')} giorni fa"],
        ["Manager score", f"{dip.get('manager_score', 'N/D')}/10",
         "eNPS score", f"{dip.get('enps_score', 'N/D')}/10"],
    ]

    formatted_rows = []
    for r in info_rows:
        formatted_rows.append([
            Paragraph(r[0], styles["Label"]),
            Paragraph(str(r[1]), styles["Value"]),
            Paragraph(r[2], styles["Label"]),
            Paragraph(str(r[3]), styles["Value"]),
        ])

    info_table = Table(formatted_rows, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5.5*cm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), C_LIGHT_GRAY),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, C_BORDER),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [C_LIGHT_GRAY, colors.white]),
    ]))
    elements.append(info_table)

    # Sezione analisi LLM
    sintesi = analysis.get("sintesi", "—")
    azione_imm = analysis.get("azione_immediata", "—")
    azione_mt = analysis.get("azione_medio_termine", "—")
    segnali = analysis.get("segnali_positivi", [])
    urgenza = analysis.get("urgenza", "—")

    analysis_data = [
        [Paragraph("<b>ANALISI</b>", styles["Label"]),
         Paragraph(sintesi, styles["CardBody"])],
        [Paragraph("<b>AZIONE IMMEDIATA</b>", styles["Label"]),
         Paragraph(azione_imm, styles["CardBody"])],
        [Paragraph("<b>MEDIO TERMINE</b>", styles["Label"]),
         Paragraph(azione_mt, styles["CardBody"])],
    ]
    if segnali:
        analysis_data.append([
            Paragraph("<b>PUNTI DI FORZA</b>", styles["Label"]),
            Paragraph(", ".join(segnali), styles["CardBody"]),
        ])
    analysis_data.append([
        Paragraph("<b>URGENZA</b>", styles["Label"]),
        Paragraph(f"<b>{urgenza.upper()}</b>", ParagraphStyle(
            "Urgenza", parent=styles["CardBody"],
            textColor=risk_color, fontName="Helvetica-Bold"
        )),
    ])

    analysis_table = Table(analysis_data, colWidths=[3.5*cm, 14*cm])
    analysis_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, C_BORDER),
        ("BOX", (0, 0), (-1, -1), 0.5, C_BORDER),
    ]))
    elements.append(analysis_table)
    elements.append(Spacer(1, 0.6*cm))
    return elements


def generate_pdf(
    df_scored: pd.DataFrame,
    llm_results: list,
    output_path: str = "reports/hr_churn_report.pdf",
    company_name: str = "Azienda",
) -> str:
    """
    Genera il report PDF completo.

    Args:
        df_scored: DataFrame con tutti i dipendenti scorati
        llm_results: lista di dict {employee, analysis} da llm_engine.analyze_batch()
        output_path: percorso di output del PDF
        company_name: nome azienda per intestazione

    Returns:
        percorso del file generato
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    styles = _build_styles()
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.8*cm, bottomMargin=1.8*cm,
    )

    elements = []
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── COPERTINA ──
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph("HR Churn Report", styles["Title2"]))
    elements.append(Paragraph(f"{company_name}  ·  Generato il {now}", styles["Subtitle"]))
    elements.append(HRFlowable(width="100%", thickness=2, color=C_ACCENT, spaceAfter=16))

    # ── RIEPILOGO ──
    elements += _summary_table(df_scored, styles)
    elements.append(PageBreak())

    # ── SCHEDE DIPENDENTI ──
    elements.append(Paragraph(
        f"Analisi dettagliata — Top {len(llm_results)} dipendenti a rischio",
        styles["SectionHeader"]
    ))
    elements.append(Spacer(1, 0.3*cm))

    for i, result in enumerate(llm_results, start=1):
        elements += _employee_card(result["employee"], result["analysis"], styles, i)
        # Page break ogni 2 dipendenti per leggibilità
        if i % 2 == 0 and i < len(llm_results):
            elements.append(PageBreak())

    # ── FOOTER PAGE ──
    elements.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph(
        f"Report generato da HR Churn Predictor  ·  {now}  ·  Dati riservati ad uso interno",
        styles["Footer"]
    ))

    doc.build(elements)
    print(f"PDF generato: {output_path}")
    return output_path


# ─────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────

def _xl_header_style(ws, row: int, headers: list, col_widths: list):
    """Applica stile header a una riga Excel."""
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1A1A2E")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 28


def generate_excel(
    df_scored: pd.DataFrame,
    llm_results: list,
    output_path: str = "reports/hr_churn_report.xlsx",
    company_name: str = "Azienda",
) -> str:
    """
    Genera il report Excel con 3 fogli:
    - Dashboard: KPI e grafico per dipartimento
    - Classifica: tutti i dipendenti con score e causa
    - Analisi LLM: suggerimenti dettagliati per i top N
    """
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb = openpyxl.Workbook()

    _build_dashboard_sheet(wb, df_scored, company_name)
    _build_classifica_sheet(wb, df_scored)
    _build_llm_sheet(wb, llm_results)

    # Rimuovi foglio default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"Excel generato: {output_path}")
    return output_path


def _build_dashboard_sheet(wb, df_scored: pd.DataFrame, company_name: str):
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    # Titolo
    ws.merge_cells("B2:H2")
    title_cell = ws["B2"]
    title_cell.value = f"HR Churn Report — {company_name}"
    title_cell.font = Font(bold=True, size=16, color="1A1A2E")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:H3")
    sub_cell = ws["B3"]
    sub_cell.value = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  {len(df_scored)} dipendenti analizzati"
    sub_cell.font = Font(size=10, color="64748B")
    ws.row_dimensions[3].height = 20

    # KPI boxes (riga 5)
    risk_counts = df_scored["risk_level"].value_counts()
    kpis = [
        ("Totale analizzati", len(df_scored), "1A1A2E", "FFFFFF"),
        ("🔴 Critici",   risk_counts.get("Critico", 0), "DC2626", "FFFFFF"),
        ("🟠 Alti",      risk_counts.get("Alto", 0),    "EA580C", "FFFFFF"),
        ("🟡 Medi",      risk_counts.get("Medio", 0),   "CA8A04", "FFFFFF"),
        ("🟢 Bassi",     risk_counts.get("Basso", 0),   "16A34A", "FFFFFF"),
        ("Score medio",  f"{df_scored['churn_score'].mean():.1f}", "334155", "FFFFFF"),
    ]

    col_map = ["B", "C", "D", "E", "F", "G"]
    for i, (label, value, bg, fg) in enumerate(kpis):
        col = col_map[i]
        ws.merge_cells(f"{col}5:{col}5")
        ws.merge_cells(f"{col}6:{col}6")
        label_cell = ws[f"{col}5"]
        label_cell.value = label
        label_cell.font = Font(size=8, color=bg, bold=True)
        label_cell.alignment = Alignment(horizontal="center")
        value_cell = ws[f"{col}6"]
        value_cell.value = value
        value_cell.font = Font(size=22, bold=True, color=bg)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[6].height = 40
        ws.column_dimensions[col].width = 14

    # Grafico a barre: rischio per dipartimento
    dept_risk = df_scored[df_scored["risk_level"].isin(["Critico", "Alto"])].groupby(
        "dipartimento"
    ).size().reset_index(name="count").sort_values("count", ascending=False)

    chart_start_row = 9
    ws.cell(row=chart_start_row, column=2, value="Dipartimento").font = Font(bold=True, size=9)
    ws.cell(row=chart_start_row, column=3, value="Dipendenti a rischio").font = Font(bold=True, size=9)

    for i, (_, row_data) in enumerate(dept_risk.iterrows(), start=1):
        ws.cell(row=chart_start_row + i, column=2, value=row_data["dipartimento"])
        ws.cell(row=chart_start_row + i, column=3, value=int(row_data["count"]))

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Dipendenti Critici/Alti per Dipartimento"
    chart.y_axis.title = ""
    chart.x_axis.title = "Dipendenti a rischio"
    chart.style = 10
    chart.width = 14
    chart.height = 10

    data_ref = Reference(ws, min_col=3, min_row=chart_start_row,
                         max_row=chart_start_row + len(dept_risk))
    cats_ref = Reference(ws, min_col=2, min_row=chart_start_row + 1,
                         max_row=chart_start_row + len(dept_risk))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "E94560"

    ws.add_chart(chart, "E8")


def _build_classifica_sheet(wb, df_scored: pd.DataFrame):
    ws = wb.create_sheet("📋 Classifica")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    columns = [
        ("Rank",                        6),
        ("Nome",                       12),
        ("Cognome",                    14),
        ("Dipartimento",               14),
        ("Ruolo",                      18),
        ("Score",                       8),
        ("Rischio",                    10),
        ("Causa Principale",           24),
        ("Età",                         6),
        ("Anni in azienda",            10),
        ("Stipendio",                  12),
        ("Ore/sett",                    9),
        ("Ultimo HR check-in (gg)",    14),
        ("Manager score",              10),
        ("eNPS",                        8),
    ]
    headers = [c[0] for c in columns]
    widths  = [c[1] for c in columns]
    _xl_header_style(ws, 1, headers, widths)

    risk_fills = {
        "Critico": PatternFill("solid", fgColor="FEE2E2"),
        "Alto":    PatternFill("solid", fgColor="FFEDD5"),
        "Medio":   PatternFill("solid", fgColor="FEF9C3"),
        "Basso":   PatternFill("solid", fgColor="DCFCE7"),
    }
    risk_fonts = {
        "Critico": Font(bold=True, color="DC2626", size=9),
        "Alto":    Font(bold=True, color="EA580C", size=9),
        "Medio":   Font(bold=True, color="CA8A04", size=9),
        "Basso":   Font(bold=True, color="16A34A", size=9),
    }

    for rank, (_, row) in enumerate(df_scored.iterrows(), start=1):
        r = rank + 1
        risk = str(row.get("risk_level", "Basso"))
        row_fill = risk_fills.get(risk, PatternFill())
        row_font = Font(size=9)

        values = [
            rank,
            row.get("nome", ""),
            row.get("cognome", ""),
            row.get("dipartimento", ""),
            row.get("ruolo", ""),
            round(float(row.get("churn_score", 0)), 1),
            risk,
            row.get("causa_principale", ""),
            row.get("eta", ""),
            round(float(row.get("anni_in_azienda", 0)), 1),
            row.get("stipendio_annuo", ""),
            round(float(row.get("ore_settimanali_medie", 0)), 1),
            row.get("giorni_ultimo_checkin_hr", ""),
            round(float(row.get("manager_score", 0)), 1),
            round(float(row.get("enps_score", 0)), 1),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = risk_fonts.get(risk, row_font) if col_idx == 7 else row_font
            cell.fill = row_fill if col_idx == 7 else PatternFill()
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 6, 7, 9, 10, 12, 13, 14, 15) else "left",
                                       vertical="center")
            if col_idx == 11:  # Stipendio
                cell.number_format = "€#,##0"

    # Color scale su colonna Score (F)
    score_range = f"F2:F{len(df_scored)+1}"
    ws.conditional_formatting.add(score_range, ColorScaleRule(
        start_type="num", start_value=0,   start_color="16A34A",
        mid_type="num",   mid_value=50,    mid_color="CA8A04",
        end_type="num",   end_value=100,   end_color="DC2626",
    ))


def _build_llm_sheet(wb, llm_results: list):
    ws = wb.create_sheet("🤖 Analisi AI")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    columns = [
        ("Nome",                10),
        ("Cognome",             14),
        ("Score",                8),
        ("Rischio",             10),
        ("Causa principale",    22),
        ("Analisi",             40),
        ("Azione immediata",    36),
        ("Azione medio termine",36),
        ("Punti di forza",      28),
        ("Urgenza",             12),
    ]
    headers = [c[0] for c in columns]
    widths  = [c[1] for c in columns]
    _xl_header_style(ws, 1, headers, widths)

    urgenza_fills = {
        "immediata": PatternFill("solid", fgColor="FEE2E2"),
        "alta":      PatternFill("solid", fgColor="FFEDD5"),
        "media":     PatternFill("solid", fgColor="FEF9C3"),
        "bassa":     PatternFill("solid", fgColor="DCFCE7"),
    }

    for i, result in enumerate(llm_results, start=2):
        emp = result["employee"]
        ana = result["analysis"]
        risk = str(emp.get("risk_level", ""))
        urgenza = str(ana.get("urgenza", "media")).lower()

        values = [
            emp.get("nome", ""),
            emp.get("cognome", ""),
            round(float(emp.get("churn_score", 0)), 1),
            risk,
            ana.get("causa_principale", ""),
            ana.get("sintesi", ""),
            ana.get("azione_immediata", ""),
            ana.get("azione_medio_termine", ""),
            ", ".join(ana.get("segnali_positivi", [])),
            urgenza.upper(),
        ]

        row_fill = urgenza_fills.get(urgenza, PatternFill())

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center" if col_idx in (3, 4, 10) else "left"
            )
            cell.font = Font(size=9)
            if col_idx == 10:
                cell.fill = row_fill
                cell.font = Font(size=9, bold=True)

        ws.row_dimensions[i].height = 55


# ─────────────────────────────────────────────
# COLAB DOWNLOAD HELPER
# ─────────────────────────────────────────────

def download_in_colab(file_path: str):
    """Su Google Colab, avvia il download automatico del file."""
    try:
        from google.colab import files
        files.download(file_path)
        print(f"Download avviato: {file_path}")
    except ImportError:
        print(f"File disponibile in: {file_path}")
